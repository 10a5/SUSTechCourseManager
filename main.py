import numpy as np
from collections import defaultdict
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QHBoxLayout, QVBoxLayout,
                           QLineEdit, QPushButton, QLabel, QFrame, QScrollArea, QGroupBox,
                           QTextEdit, QMessageBox, QTableWidget, QTableWidgetItem, QHeaderView,
                           QListWidget, QListWidgetItem, QDialog, QDialogButtonBox)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor, QBrush
import requests
import os
import sys
import getpass
import re
from re import findall
from json import loads
import pandas as pd
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import warnings
from PyQt5.QtWidgets import QFileDialog
from urllib3.exceptions import InsecureRequestWarning

os.chdir(os.path.dirname(os.path.abspath(__file__)))

course_reference = {
    "bxxk": "通识必修选课",
    "xxxk": "通识选修选课",
    "kzyxk": "培养方案内选课",
    "zynknjxk": "非培养方案内选课",
    "jhnxk": "重修选课"
}
def warn(message, category, filename, lineno, file=None, line=None):
    if category is not InsecureRequestWarning:
        sys.stderr.write(warnings.formatwarning(message, category, filename, lineno, line))
warnings.showwarning = warn

class CourseSchedulerApp(QMainWindow):
    def __init__(self, course_list):
        super().__init__()
        self.setWindowTitle("课程表管理系统")
        self.setGeometry(100, 100, 1800, 800)
        self.course_list = course_list
        self.filtered_courses = []
        self.schedules = []
        self.current_schedule_idx = -1
        self.locked_time_slots = np.zeros((16, 7, 11), dtype=bool)
        
        self.init_ui()
        
        self.show()
        self.setWindowState(self.windowState() & ~Qt.WindowMinimized | Qt.WindowActive)
        self.raise_()
        for _ in range(3):
            self.activateWindow()
            QApplication.processEvents()
    
    def init_ui(self):
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(10)
        
        left_panel = QFrame()
        left_panel.setFrameShape(QFrame.StyledPanel)
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(10, 10, 10, 10)
        
        search_group = QGroupBox("课程搜索与选择")
        search_layout = QVBoxLayout(search_group)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("输入课程关键词")
        self.search_btn = QPushButton("搜索")
        
        self.search_result_list = QListWidget()
        self.search_result_list.setSelectionMode(QListWidget.MultiSelection)
        
        self.selected_courses_list = QListWidget()
        
        btn_frame = QFrame()
        btn_layout = QHBoxLayout(btn_frame)
        self.add_selected_btn = QPushButton("添加选中课程")
        self.remove_selected_btn = QPushButton("移除选中课程")
        self.clear_selected_btn = QPushButton("清空已选")
        btn_layout.addWidget(self.add_selected_btn)
        btn_layout.addWidget(self.remove_selected_btn)
        btn_layout.addWidget(self.clear_selected_btn)
        
        search_layout.addWidget(QLabel("搜索课程:"))
        search_layout.addWidget(self.search_input)
        search_layout.addWidget(self.search_btn)
        search_layout.addWidget(QLabel("搜索结果:"))
        search_layout.addWidget(self.search_result_list)
        search_layout.addWidget(btn_frame)
        search_layout.addWidget(QLabel("已选课程:"))
        search_layout.addWidget(self.selected_courses_list)
        
        left_layout.addWidget(search_group)
        
        center_panel = QFrame()
        center_panel.setFrameShape(QFrame.StyledPanel)
        center_layout = QVBoxLayout(center_panel)
        center_layout.setContentsMargins(10, 10, 10, 10)
        
        generate_group = QGroupBox("课程表生成")
        generate_layout = QVBoxLayout(generate_group)

        self.export_btn = QPushButton("导出当前课表到Excel")
        self.export_btn.setEnabled(False)
        generate_layout.addWidget(self.export_btn)
        
        self.generate_btn = QPushButton("生成所有可能的课程表")
        generate_layout.addWidget(self.generate_btn)
        
        self.schedule_count_label = QLabel("找到 0 个有效课程表")
        generate_layout.addWidget(self.schedule_count_label)
        
        nav_frame = QFrame()
        nav_layout = QHBoxLayout(nav_frame)
        
        self.prev_btn = QPushButton("上一个")
        self.prev_btn.setEnabled(False)
        self.next_btn = QPushButton("下一个")
        self.next_btn.setEnabled(False)
        
        nav_layout.addWidget(self.prev_btn)
        nav_layout.addWidget(self.next_btn)
        generate_layout.addWidget(nav_frame)
        
        lock_frame = QFrame()
        lock_layout = QHBoxLayout(lock_frame)
        self.clear_locks_btn = QPushButton("清除所有时间锁定")
        self.show_locks_btn = QPushButton("显示锁定时间段（双击表格单元格锁定/解锁）")
        lock_layout.addWidget(self.clear_locks_btn)
        lock_layout.addWidget(self.show_locks_btn)
        generate_layout.addWidget(lock_frame)
        
        center_layout.addWidget(generate_group)
        
        schedule_group = QGroupBox("课程表展示")
        schedule_layout = QVBoxLayout(schedule_group)
        
        self.schedule_table = QTableWidget()
        self.schedule_table.setColumnCount(7)
        self.schedule_table.setHorizontalHeaderLabels(["周一", "周二", "周三", "周四", "周五", "周六", "周日"])
        self.schedule_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.schedule_table.setRowCount(11)
        self.schedule_table.setVerticalHeaderLabels([f"第{i+1}节" for i in range(11)])
        self.schedule_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)
        
        self.schedule_table.cellDoubleClicked.connect(self.toggle_time_slot_lock)
        
        schedule_layout.addWidget(self.schedule_table)
        center_layout.addWidget(schedule_group)
        
        right_panel = QFrame()
        right_panel.setFrameShape(QFrame.StyledPanel)
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(10, 10, 10, 10)
        
        available_group = QGroupBox("可用课程查询")
        available_layout = QVBoxLayout(available_group)
        
        self.available_search_input = QLineEdit()
        self.available_search_input.setPlaceholderText("输入课程名、教师或时间...")
        available_layout.addWidget(self.available_search_input)
        
        self.available_search_btn = QPushButton("查询可用课程")
        available_layout.addWidget(self.available_search_btn)
        
        self.available_courses_list = QListWidget()
        available_layout.addWidget(self.available_courses_list)
        
        self.add_to_schedule_btn = QPushButton("加入当前课表")
        available_layout.addWidget(self.add_to_schedule_btn)
        
        right_layout.addWidget(available_group)
        
        main_layout.addWidget(left_panel, 1)
        main_layout.addWidget(center_panel, 4)
        main_layout.addWidget(right_panel, 1)
        
        self.search_btn.clicked.connect(self.search_courses)
        self.add_selected_btn.clicked.connect(self.add_selected_courses)
        self.remove_selected_btn.clicked.connect(self.remove_selected_courses)
        self.clear_selected_btn.clicked.connect(self.clear_selected_courses)
        self.generate_btn.clicked.connect(self.generate_schedules)
        self.prev_btn.clicked.connect(self.show_prev_schedule)
        self.next_btn.clicked.connect(self.show_next_schedule)
        self.available_search_btn.clicked.connect(self.search_available_courses)
        self.add_to_schedule_btn.clicked.connect(self.add_selected_available_course)
        self.export_btn.clicked.connect(self.export_to_excel)
        self.clear_locks_btn.clicked.connect(self.clear_all_locks)
        self.show_locks_btn.clicked.connect(self.show_locked_time_slots)
    
    def toggle_time_slot_lock(self, row, column):
        day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        time_slot = f"第{row+1}节"
        day_name = day_names[column]
        is_currently_locked = np.any(self.locked_time_slots[:, column, row])
        if is_currently_locked:
            for week in range(16):
                self.locked_time_slots[week, column, row] = False
            self.update_table_appearance()
            QMessageBox.information(self, "时间段解锁", f"已解锁 {day_name} {time_slot}")
        else:
            for week in range(16):
                self.locked_time_slots[week, column, row] = True
            self.update_table_appearance()
            QMessageBox.information(self, "时间段锁定", f"已锁定 {day_name} {time_slot}")
    
    def update_table_appearance(self):
        for day in range(7):
            for time in range(11):
                item = self.schedule_table.item(time, day)
                if item is None:
                    item = QTableWidgetItem()
                    self.schedule_table.setItem(time, day, item)
                
                if np.any(self.locked_time_slots[:, day, time]):
                    item.setBackground(QBrush(QColor(255, 200, 200)))
                    item.setToolTip("该时间段已被锁定")
                else:
                    item.setBackground(QBrush(QColor(255, 255, 255)))  
                    item.setToolTip("")
    
    def clear_all_locks(self):
        reply = QMessageBox.question(
            self, 
            "确认清除",
            "确定要清除所有时间锁定吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.locked_time_slots = np.zeros((16, 7, 11), dtype=bool)
            self.update_table_appearance()
            QMessageBox.information(self, "成功", "已清除所有时间锁定")
    
    def show_locked_time_slots(self):
        locked_slots = []
        day_names = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
        
        for day in range(7):
            for time in range(11):
                if np.any(self.locked_time_slots[:, day, time]):
                    locked_slots.append(f"{day_names[day]} 第{time+1}节")
        
        if not locked_slots:
            QMessageBox.information(self, "锁定时间段", "当前没有锁定的时间段")
        else:
            message = "当前锁定的时间段:\n\n" + "\n".join(locked_slots)
            QMessageBox.information(self, "锁定时间段", message)
    
    def search_courses(self):
        keyword = self.search_input.text().strip()
        if not keyword:
            QMessageBox.warning(self, "提示", "请输入搜索关键词")
            return
        
        self.search_result_list.clear()
        
        matched_courses = set()
        for course in self.course_list:
            if keyword.lower() in course[0].lower():
                matched_courses.add(course[0])
        
        if not matched_courses:
            QMessageBox.information(self, "提示", f"未找到包含'{keyword}'的课程")
            return
        
        for course_name in sorted(matched_courses):
            self.search_result_list.addItem(course_name)
    
    def add_selected_courses(self):
        selected_items = self.search_result_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "提示", "请先在搜索结果中选择课程")
            return
        
        for item in selected_items:
            course_name = item.text()
            courses = [c for c in self.course_list if c[0] == course_name]
            
            if len(courses) == 1:
                self.add_course_to_selected(courses[0])
            else:
                self.show_course_selection_dialog(course_name, courses)

    def remove_selected_courses(self):
        selected_items = self.selected_courses_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "提示", "请先选择要移除的课程")
            return
        
        reply = QMessageBox.question(
            self, 
            "确认移除",
            f"确定要移除这 {len(selected_items)} 门课程吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            for item in selected_items:
                course = item.data(Qt.UserRole)
                row = self.selected_courses_list.row(item)
                self.selected_courses_list.takeItem(row)
                
                for i, c in enumerate(self.filtered_courses):
                    if (c[0] == course[0] and 
                        c[1] == course[1] and 
                        np.array_equal(c[2], course[2])):
                        self.filtered_courses.pop(i)
                        break
    
    def show_course_selection_dialog(self, course_name, courses):
        dialog = QDialog(self)
        dialog.setWindowTitle(f"选择 {course_name} 纳入考虑的课程")
        dialog.setMinimumWidth(500) 
        
        layout = QVBoxLayout()

        list_widget = QListWidget()
        for course in courses:
            item_text = f"{course[1]}\n时间: {course[3]}" 
            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, course)
            list_widget.addItem(item)
        
        list_widget.setSelectionMode(QListWidget.MultiSelection)
        
        btn_frame_top = QFrame()
        btn_layout_top = QHBoxLayout(btn_frame_top)
        
        select_all_btn = QPushButton("全选")
        select_all_btn.clicked.connect(lambda: list_widget.selectAll())
        
        clear_selection_btn = QPushButton("清空选择")
        clear_selection_btn.clicked.connect(lambda: list_widget.clearSelection())
        
        btn_layout_top.addWidget(select_all_btn)
        btn_layout_top.addWidget(clear_selection_btn)
        btn_layout_top.addStretch()
        
        btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btn_box.accepted.connect(dialog.accept)
        btn_box.rejected.connect(dialog.reject)
        
        layout.addWidget(btn_frame_top)
        layout.addWidget(list_widget)
        layout.addWidget(btn_box)
        
        dialog.setLayout(layout)
        
        if dialog.exec_() == QDialog.Accepted:
            for item in list_widget.selectedItems():
                course = item.data(Qt.UserRole)
                self.add_course_to_selected(course)
    
    def add_course_to_selected(self, course):
        if np.any(np.logical_and(self.locked_time_slots, course[2])):
            QMessageBox.warning(self, "时间冲突", 
                               f"课程 '{course[0]}' 与锁定的时间段冲突，无法添加")
            return
        
        for i in range(self.selected_courses_list.count()):
            existing_item = self.selected_courses_list.item(i)
            existing_course = existing_item.data(Qt.UserRole)
            if (existing_course[0] == course[0] and 
                existing_course[1] == course[1] and 
                np.array_equal(existing_course[2], course[2])):
                QMessageBox.warning(self, "提示", "该课程已添加")
                return
        
        item_text = f"{course[0]} ({course[1]})\n时间: {course[3]}"
        item = QListWidgetItem(item_text)
        item.setData(Qt.UserRole, course)
        self.selected_courses_list.addItem(item)
        
        self.filtered_courses.append(course)
        
        self.update_selected_count()

    def update_selected_count(self):
        unique_course_names = set()
        for c in self.filtered_courses:
            unique_course_names.add(c[0])
        
        count = len(unique_course_names)
        self.selected_courses_list.setToolTip(f"已选课程种类: {count}/12 种")
        if count >= 12:
            self.selected_courses_list.setStyleSheet("QListWidget { border: 2px solid red; }")
        else:
            self.selected_courses_list.setStyleSheet("")
    
    def clear_selected_courses(self):
        self.selected_courses_list.clear()
        self.filtered_courses = []
    
    def generate_schedules(self):
        if not self.filtered_courses:
            QMessageBox.warning(self, "警告", "请先添加课程")
            return
        
        course_groups = defaultdict(list)
        for course in self.filtered_courses:
            course_groups[course[0]].append(course) 
        
        self.schedules = []
        max_courses = 0
        
        def find_max_courses(group_keys, index, current_count, combined_schedule):
            nonlocal max_courses
            
            if index >= len(group_keys):
                if current_count > max_courses:
                    max_courses = current_count
                return
            
            find_max_courses(group_keys, index + 1, current_count, combined_schedule)
            
            current_group = course_groups[group_keys[index]]
            for course in current_group:
                if (not np.any(np.logical_and(combined_schedule[0], course[2][0])) and \
                   not np.any(np.logical_and(self.locked_time_slots, course[2]))):
                    new_schedule = np.logical_or(combined_schedule, course[2])
                    find_max_courses(group_keys, index + 1, current_count + 1, new_schedule)
        
        group_keys = list(course_groups.keys())
        find_max_courses(group_keys, 0, 0, np.zeros((16, 7, 11), dtype=bool))
        
        if max_courses == 0:
            QMessageBox.warning(self, "提示", "没有找到有效的课程表组合（可能与锁定的时间段冲突）")
            return
        
        current_schedule = []
        combined_schedule = np.zeros((16, 7, 11), dtype=bool)
        
        def backtrack_maximal(group_keys, index):
            nonlocal combined_schedule
            
            if len(current_schedule) == max_courses:
                schedule_key = tuple(
                    (c[0], c[1], tuple(zip(*np.where(c[2])))) 
                    for c in sorted(current_schedule, key=lambda x: x[0])
                )
                if schedule_key not in self.schedules:
                    self.schedules.append([course for course in current_schedule])
                return
            
            if index >= len(group_keys):
                return
            
            backtrack_maximal(group_keys, index + 1)
            
            current_group = course_groups[group_keys[index]]
            for course in current_group:
                if (not np.any(np.logical_and(combined_schedule[0], course[2][0])) and \
                   not np.any(np.logical_and(self.locked_time_slots, course[2]))):
                    current_schedule.append(course)
                    old_schedule = combined_schedule.copy()
                    combined_schedule = np.logical_or(combined_schedule, course[2])
                    
                    backtrack_maximal(group_keys, index + 1)
                    
                    current_schedule.pop()
                    combined_schedule = old_schedule
        
        backtrack_maximal(group_keys, 0)
        
        self.schedule_count_label.setText(f"找到 {len(self.schedules)} 个包含 {max_courses} 门课程的有效课程表")
        
        if self.schedules:
            self.current_schedule_idx = 0
            self.show_schedule(self.current_schedule_idx)
            self.update_nav_buttons()
            
            detail = f"找到 {len(self.schedules)} 个包含 {max_courses} 门课程的课程表:\n"
            for i, course in enumerate(self.schedules[0]):
                detail += f"{i+1}. {course[0]} ({course[1]})\n"
            
            QMessageBox.information(self, "提示", detail)
        else:
            QMessageBox.warning(self, "提示", "没有找到有效的课程表组合（可能与锁定的时间段冲突）")
    
    def show_schedule(self, idx):
        if 0 <= idx < len(self.schedules):
            self.export_btn.setEnabled(True)
            schedule = self.schedules[idx]
            
            weekly_schedule = defaultdict(lambda: defaultdict(list))
            
            for course in schedule:
                for week in range(16):
                    for day in range(7):
                        for time in range(11):
                            if course[2][week, day, time]:
                                course_info = f"{course[0]}({course[1]})"
                                weekly_schedule[day][time].append(course_info)
            
            self.schedule_table.clearContents()
            for day in range(7):
                for time in range(11):
                    courses_in_slot = weekly_schedule[day][time]
                    if courses_in_slot:
                        item = QTableWidgetItem()
                        item.setTextAlignment(Qt.AlignCenter)
                        
                        unique_courses = sorted(set(courses_in_slot))
                        item.setText("\n".join(unique_courses))
                        
                        if len(unique_courses) > 1:
                            item.setBackground(QBrush(QColor(255, 220, 200)))
                        else:
                            item.setBackground(QBrush(QColor(200, 255, 200))) 
                        
                        self.schedule_table.setItem(time, day, item)
            
            self.update_table_appearance()
        else:
            self.export_btn.setEnabled(False)
    
    def show_prev_schedule(self):
        if self.current_schedule_idx > 0:
            self.current_schedule_idx -= 1
            self.show_schedule(self.current_schedule_idx)
            self.update_nav_buttons()
    
    def show_next_schedule(self):
        if self.current_schedule_idx < len(self.schedules) - 1:
            self.current_schedule_idx += 1
            self.show_schedule(self.current_schedule_idx)
            self.update_nav_buttons()
    
    def update_nav_buttons(self):
        self.prev_btn.setEnabled(self.current_schedule_idx > 0)
        self.next_btn.setEnabled(self.current_schedule_idx < len(self.schedules) - 1)
    
    def search_available_courses(self):
        query = self.available_search_input.text().strip().lower()
        self.available_courses_list.clear()
        
        if not self.schedules or self.current_schedule_idx == -1:
            QMessageBox.warning(self, "警告", "请先创建或选择一个课表！")
            return
        
        current_schedule = self.schedules[self.current_schedule_idx]
        existing_course_names = {course[0] for course in current_schedule}
        booked_slots = np.zeros((16, 7, 11), dtype=bool)
        for course in current_schedule:
            booked_slots = np.logical_or(booked_slots, course[2])

        for course in self.course_list:
            if course[0] in existing_course_names:
                continue
            
            matches_query = (
                query in course[0].lower() or
                query in course[1].lower() or
                query in course[3].lower()
            )
            
            conflicts_with_schedule = np.any(np.logical_and(booked_slots, course[2]))
            conflicts_with_locks = np.any(np.logical_and(self.locked_time_slots, course[2]))
            
            if matches_query and not conflicts_with_schedule and not conflicts_with_locks:
                item = QListWidgetItem(f"{course[0]} - {course[1]}\n时间: {course[3]}")
                item.setData(Qt.UserRole, course)
                self.available_courses_list.addItem(item)

    def add_selected_available_course(self):
        selected_item = self.available_courses_list.currentItem()
        if not selected_item:
            QMessageBox.warning(self, "警告", "请先选择一个课程！")
            return
        
        if not self.schedules or self.current_schedule_idx == -1:
            QMessageBox.warning(self, "警告", "请先创建或选择一个课表！")
            return
        
        course = selected_item.data(Qt.UserRole)
        current_schedule = self.schedules[self.current_schedule_idx]
        
        if any(c[0] == course[0] for c in current_schedule):
            QMessageBox.warning(self, "警告", f"课程'{course[0]}'已存在于当前课表中！")
            return
        
        if np.any(np.logical_and(self.locked_time_slots, course[2])):
            QMessageBox.warning(self, "时间冲突", 
                               f"课程 '{course[0]}' 与锁定的时间段冲突，无法添加")
            return
        
        current_schedule.append(course)
        
        self.show_schedule(self.current_schedule_idx)
        QMessageBox.information(self, "成功", f"已添加课程: {course[0]}({course[1]})")
        
        self.search_available_courses()

    def export_to_excel(self):
        if not hasattr(self, 'schedules') or not self.schedules or self.current_schedule_idx == -1:
            QMessageBox.warning(self, "警告", "没有可导出的课程表")
            return
        
        try:
            options = QFileDialog.Options()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"课程表_{timestamp}.xlsx"
            
            filename, _ = QFileDialog.getSaveFileName(
                self, 
                "选择保存位置", 
                default_filename, 
                "Excel文件 (*.xlsx)", 
                options=options
            )
            
            if not filename:
                return
            
            if not filename.endswith('.xlsx'):
                filename += '.xlsx'
            
            schedule = self.schedules[self.current_schedule_idx]
            weekdays = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            periods = [f"第{i+1}节" for i in range(11)]
            timetable = [["" for _ in range(7)] for _ in range(11)]
            
            for course in schedule:
                course_name = f"{course[0]}\n({course[1]})"
                for day in range(7):
                    for period in range(11):
                        if course[2][1, day, period]:
                            if timetable[period][day]:
                                timetable[period][day] += "\n" + course_name
                            else:
                                timetable[period][day] = course_name

            df = pd.DataFrame(timetable, columns=weekdays, index=periods)
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=True, sheet_name='课程表')
                
                worksheet = writer.sheets['课程表']
                worksheet.column_dimensions['A'].width = 8
                for col in range(1, 8): 
                    column_letter = get_column_letter(col + 1)
                    worksheet.column_dimensions[column_letter].width = 25 
                    
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            QMessageBox.information(self, "导出成功", f"课程表已成功导出到:\n{filename}")
        
        except PermissionError:
            QMessageBox.critical(self, "权限错误", 
                            f"没有权限写入文件:\n{filename}\n\n请尝试选择其他位置或关闭已打开的文件。")
        except ImportError:
            QMessageBox.critical(self, "错误", "导出失败: 请先安装pandas和openpyxl库\n\n在命令行运行:\npip install pandas openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出过程中发生错误:\n{str(e)}")

def parse_schedule(kcxx_text):
    schedule = np.zeros((16, 7, 11), dtype=bool)
    time_texts = []
    
    schedule_patterns = re.findall(r'<p>([^<]+?)</p>', kcxx_text)
    for pattern in schedule_patterns:
        if "周," not in pattern:
            continue
        if not any(str(i) in pattern for i in range(1, 12)): 
            continue
        
        time_texts.append(pattern) 
        
        week_part = pattern.split('周,')[0].strip()
        weeks = parse_weeks(week_part)
        day_period_part = pattern.split('周,')[1].strip()
        day, periods = parse_day_periods(day_period_part)
        
        for week in weeks:
            for period in periods:
                if 1 <= week <= 16 and 1 <= day <= 7 and 1 <= period <= 11:
                    schedule[week-1, day-1, period-1] = True
    
    time_text = "；".join(time_texts)
    return schedule, time_text

def parse_weeks(week_str):
    weeks = []
    week_str = week_str.strip().replace('周', '').replace(' ', '')
    if not week_str:
        return weeks
    parts = week_str.split(',')
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if '单' in part or '双' in part:
            parity = '单' if '单' in part else '双'
            num_part = part.replace(parity, '')
            if '-' in num_part:
                start, end = map(int, num_part.split('-'))
                weeks.extend([
                    w for w in range(start, end + 1) 
                    if (w % 2 == 1 if parity == '单' else w % 2 == 0)
                ])
            else:
                num = int(num_part)
                if (num % 2 == 1 if parity == '单' else num % 2 == 0):
                    weeks.append(num)
        elif '-' in part:
            start, end = map(int, part.split('-'))
            weeks.extend(range(start, end + 1))
        else:
            try:
                weeks.append(int(part))
            except ValueError:
                continue
    weeks = sorted(list(set(weeks)))
    return weeks

def parse_day_periods(day_period_str):
    day_map = {
        '星期一': 1, '星期二': 2, '星期三': 3, '星期四': 4,
        '星期五': 5, '星期六': 6, '星期日': 7, '星期天': 7
    }
    day = None
    for ch_day, num in day_map.items():
        if ch_day in day_period_str:
            day = num
            break
    if day is None:
        return None, []
    period_part = day_period_str.split('第')[-1]
    period_numbers = []
    current_number = ''
    for char in period_part:
        if char.isdigit():
            current_number += char
        elif current_number:
            period_numbers.append(current_number)
            current_number = ''
    if current_number:
        period_numbers.append(current_number)
    if not period_numbers:
        return day, []
    if '-' in period_part:
        if len(period_numbers) >= 2:
            start = int(period_numbers[0])
            end = int(period_numbers[1])
            periods = list(range(start, end+1))
        else:
            periods = []
    else:
        periods = [int(period_numbers[0])]
    return day, periods

def get_course(semester_data, header):
    print("正在获取课程数据...")
    class_data = []
    for course_type in ['bxxk', 'xxxk', 'kzyxk', 'zynknjxk','jhnxk']:
        data = {
                "p_xn": semester_data['p_xn'],
                "p_xq": semester_data['p_xq'],
                "p_xnxq": semester_data['p_xnxq'],
                "p_pylx": 1,
                "mxpylx": 1,
                "p_xkfsdm": course_type,
                "pageNum": 1,
                "pageSize": 1000 
            }
        print(f"正在获取 {course_reference[course_type]} 课程数据...")
        req = requests.post('https://tis.sustech.edu.cn/Xsxk/queryKxrw', data=data, headers=header, verify=False)
        raw_class_data = loads(req.text)
        if raw_class_data.get('kxrwList'):
            for i in raw_class_data['kxrwList']['list']:
                schedule_array, time_text = parse_schedule(i['kcxx'])
                print(f"获取课程: {i['kcmc']} - 助教/教师: {i['dgjsmc']}")
                class_data.append([
                    i['kcmc'],   
                    i['dgjsmc'],  
                    schedule_array,
                    time_text,
                    course_reference[course_type]
                ])
    return class_data

def login(user_name, pwd, header):
    try:
        login_url = "https://cas.sustech.edu.cn/cas/login?service=https%3A%2F%2Ftis.sustech.edu.cn%2Fcas"
        req = requests.get(login_url, verify=False)
        assert req.status_code == 200
        
        def get_execution_value(html):
            try:
                start = html.index('name="execution" value="') + len('name="execution" value="')
                end = html.index('"', start)
                return html[start:end]
            except ValueError:
                raise ValueError("无法从登录页面获取execution参数")

        data = {
            'username': user_name,
            'password': pwd,
            'execution': get_execution_value(req.text),
            '_eventId': 'submit',
        }
        
        req = requests.post(login_url, data=data, allow_redirects=False, 
                          headers=header, verify=False)
        
        if "Location" not in req.headers:
            print("登录失败，请检查用户名和密码")
            return "", ""
            
        req = requests.get(req.headers["Location"], allow_redirects=False, 
                          headers=header, verify=False)
        
        cookies = req.headers.get("Set-Cookie", "")
        route = findall(r'route=([^;]+);', cookies)[0] if 'route=' in cookies else ""
        jsessionid = findall(r'JSESSIONID=([^;]+);', cookies)[0] if 'JSESSIONID=' in cookies else ""
        
        return route, jsessionid
        
    except Exception as e:
        print(f"登录过程中发生错误: {str(e)}")
        return "", ""

if __name__ == "__main__":
    header = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.0.0 Safari/537.36",
        "x-requested-with": "XMLHttpRequest"
    }
    
    route = ""
    JSESSIONID = ""
    while route == "" or JSESSIONID == "":
        user_name = input("请输入tis账号: ")
        pwd = getpass.getpass("请输入tis密码(密码不会显示): ")
        route, JSESSIONID = login(user_name, pwd, header)
        if route == "" or JSESSIONID == "":
            print("登陆失败，请检查用户名和密码或网络连接。")
    
    header['cookie'] = f"route={route}; JSESSIONID={JSESSIONID}"
    print("登录成功！")
    
    semester_info = loads(requests.post('https://tis.sustech.edu.cn/Xsxk/queryXkdqXnxq', data={"mxpylx": 1}, headers=header, verify=False).text)
    course_list = get_course(semester_info, header)
    
    app = QApplication([])
    window = CourseSchedulerApp(course_list)
    window.show()
    app.exec_()